import openpyxl

def extract_data_from_xlsx(file_path: str) -> dict[str, list[list]]:
    """
    Extract structured data from an XLSX file.

    Args:
        file_path: Path to the XLSX file

    Returns:
        dict[str, list[list]]: Dictionary with sheet names as keys and rows as values
    """
    data = {}
    workbook = openpyxl.load_workbook(file_path, data_only=True)

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        sheet_data = []

        for row in worksheet.iter_rows(values_only=True):
            # Convert None values to empty strings and convert all to strings
            row_data = [str(cell) if cell is not None else "" for cell in row]
            # Only add rows that have at least one non-empty cell
            if any(cell.strip() for cell in row_data):
                sheet_data.append(row_data)

        data[sheet_name] = sheet_data

    workbook.close()
    return data
